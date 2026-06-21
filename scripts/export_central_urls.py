"""Export Workato Central_url_* usage to Excel (research)."""
from __future__ import annotations

import json
import re
import time
import urllib.error
import urllib.request
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("Run: pip install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
ENV_PATH = ROOT / ".env"
OUT_PATH = Path(__file__).resolve().parents[2].parent / "docs" / "workato-central-url-research.xlsx"

RECIPE_UI = "https://www.workato.com/recipes/{id}"
CONN_UI = "https://www.workato.com/connections/{id}"

CENTRAL_PROP_RE = re.compile(
    r"Central_url_(v0|v2|billing|inventories)[^'\"]*",
    re.I,
)
# URL template after account property substitution
URL_AFTER_PROP_RE = re.compile(
    r"Central_url_(?:v0|v2|billing|inventories)[^'\"]*['\"][^'\"]*\)\s*([^\"']+)",
    re.I,
)
# Broader: capture url field containing Central_url
URL_FIELD_RE = re.compile(
    r'"url"\s*:\s*"([^"]*Central_url[^"]*)"',
    re.I,
)
METHOD_RE = re.compile(r'"method"\s*:\s*"([A-Z]+)"')


def load_env() -> tuple[str, str]:
    token = base = None
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line.startswith("WORKATO_API_TOKEN="):
            token = line.split("=", 1)[1].strip()
        elif line.startswith("WORKATO_API_BASE_URL="):
            base = line.split("=", 1)[1].strip().rstrip("/")
    if not token:
        raise SystemExit("WORKATO_API_TOKEN missing in .env")
    return token, base or "https://www.workato.com/api"


def api_get(token: str, base: str, path: str, params: dict | None = None) -> object:
    url = f"{base}{path}"
    if params:
        qs = "&".join(f"{k}={urllib.request.quote(str(v))}" for k, v in params.items())
        url = f"{url}?{qs}"
    req = urllib.request.Request(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8"))


def normalize_prop(raw: str) -> str:
    m = re.match(r"(Central_url_\w+)", raw.replace("\\", ""), re.I)
    return m.group(1) if m else raw


def extract_endpoints(body: str) -> list[dict]:
    rows: list[dict] = []
    seen: set[tuple] = set()

    for m in URL_FIELD_RE.finditer(body):
        raw_url = m.group(1)
        # Unescape common JSON escapes
        raw_url = raw_url.replace("\\/", "/").replace("\\\"", '"')
        props = sorted({normalize_prop(p) for p in CENTRAL_PROP_RE.findall(raw_url)})
        if not props:
            continue
        # Method near this url (search backwards ~500 chars)
        start = max(0, m.start() - 500)
        chunk = body[start : m.end() + 200]
        method_m = METHOD_RE.search(chunk)
        method = method_m.group(1) if method_m else ""

        for prop in props:
            # Build human-readable path template
            path_template = raw_url
            for p in CENTRAL_PROP_RE.findall(raw_url):
                path_template = re.sub(
                    r"#\{_[^}]*Central_url_[^}]*\}",
                    "{" + normalize_prop("Central_url_" + p) + "}",
                    path_template,
                    count=1,
                )
            path_template = re.sub(
                r"#\{_[^}]*Central_url_[^'\"]*['\"][^'\"]*['\"][^'\"]*['\"]\s*\}",
                "{PROPERTY}",
                path_template,
            )
            # Simpler display: property + suffix after closing paren
            suffix_m = re.search(
                r"Central_url_(?:v0|v2|billing|inventories)[^)]*\)\s*([^\"']*)",
                raw_url,
                re.I,
            )
            endpoint = (prop + (suffix_m.group(1) if suffix_m else "")).replace("\\", "")

            key = (prop, method, endpoint[:200])
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "property": prop,
                    "method": method,
                    "url_template": raw_url[:500],
                    "endpoint_pattern": endpoint[:300],
                }
            )
    return rows


def try_account_properties(token: str, base: str) -> list[dict]:
    paths = ["/properties", "/account_properties", "/settings/properties"]
    for path in paths:
        try:
            data = api_get(token, base, path)
            return flatten_properties(data)
        except urllib.error.HTTPError:
            continue
    return []


def flatten_properties(data: object) -> list[dict]:
    out: list[dict] = []
    if isinstance(data, list):
        items = data
    elif isinstance(data, dict):
        items = data.get("items") or data.get("properties") or data.get("result") or []
        if not items and any(k for k in data if "central" in str(k).lower()):
            items = [data]
    else:
        return out
    for item in items:
        if not isinstance(item, dict):
            continue
        name = item.get("name") or item.get("key") or item.get("property_name") or ""
        value = item.get("value") or item.get("property_value") or item.get("default") or ""
        if "central_url" in str(name).lower() or "central_url" in str(value).lower():
            out.append({"name": name, "value": value})
    return out


def main() -> None:
    token, base = load_env()
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    print("Fetching recipes...")
    all_recipes: list[dict] = []
    page = 1
    while True:
        data = api_get(token, base, "/recipes", {"per_page": 100, "page": page})
        items = data.get("items", data) if isinstance(data, dict) else data
        if not items:
            break
        all_recipes.extend(items)
        if len(items) < 100:
            break
        page += 1
        time.sleep(0.2)

    print(f"  {len(all_recipes)} recipes")

    endpoint_rows: list[dict] = []
    recipe_rows: list[dict] = []

    for i, r in enumerate(all_recipes):
        rid = str(r.get("id", ""))
        rname = r.get("name", "")
        folder = r.get("folder_id", "")
        running = r.get("running", False)
        if (i + 1) % 25 == 0:
            print(f"  Scanning recipe {i + 1}/{len(all_recipes)}...")
        try:
            detail = api_get(token, base, f"/recipes/{rid}")
        except urllib.error.HTTPError as e:
            recipe_rows.append(
                {
                    "recipe_id": rid,
                    "recipe_name": rname,
                    "folder_id": folder,
                    "running": running,
                    "properties": "",
                    "note": f"Fetch error: {e.code}",
                }
            )
            time.sleep(0.15)
            continue

        body = json.dumps(detail) if isinstance(detail, dict) else str(detail)
        props = sorted({normalize_prop(p) for p in CENTRAL_PROP_RE.findall(body)})
        if not props:
            time.sleep(0.1)
            continue

        recipe_link = RECIPE_UI.format(id=rid)
        recipe_rows.append(
            {
                "recipe_id": rid,
                "recipe_name": rname,
                "folder_id": folder,
                "running": running,
                "properties": ", ".join(props),
                "recipe_link": recipe_link,
                "note": "",
            }
        )

        for ep in extract_endpoints(body):
            endpoint_rows.append(
                {
                    "account_property": ep["property"],
                    "http_method": ep["method"],
                    "endpoint_pattern": ep["endpoint_pattern"],
                    "url_template_snippet": ep["url_template"],
                    "recipe_id": rid,
                    "recipe_name": rname,
                    "running": running,
                    "folder_id": folder,
                    "recipe_link": recipe_link,
                }
            )
        time.sleep(0.12)

    print("Fetching connections...")
    conn_rows: list[dict] = []
    try:
        conns = api_get(token, base, "/connections", {"per_page": 100})
        conn_list = conns if isinstance(conns, list) else conns.get("items", [])
        for c in conn_list:
            name = (c.get("name") or "").lower()
            if "central" in name or "luc" in name:
                cid = c.get("id", "")
                conn_rows.append(
                    {
                        "connection_id": cid,
                        "connection_name": c.get("name", ""),
                        "application": c.get("application", ""),
                        "authorization_status": c.get("authorization_status", ""),
                        "connection_link": CONN_UI.format(id=cid),
                    }
                )
    except urllib.error.HTTPError as e:
        conn_rows.append({"connection_name": f"API error {e.code}", "connection_link": ""})

    print("Trying account properties API...")
    prop_rows = try_account_properties(token, base)

    wb = Workbook()
    # Sheet 1: Summary by property
    ws_sum = wb.active
    ws_sum.title = "Summary by Property"
    headers_sum = ["Account Property", "Recipe Count", "Example Recipe", "Recipe Link"]
    ws_sum.append(headers_sum)
    prop_counts: dict[str, list[str]] = {}
    prop_links: dict[str, str] = {}
    for rr in recipe_rows:
        for p in rr["properties"].split(", "):
            if not p:
                continue
            prop_counts.setdefault(p, []).append(rr["recipe_id"])
            prop_links.setdefault(p, rr.get("recipe_link", ""))
    for p in sorted(prop_counts.keys()):
        ids = prop_counts[p]
        link = prop_links.get(p, "")
        ws_sum.append([p, len(ids), ids[0] if ids else "", link])
        if link:
            cell = ws_sum.cell(row=ws_sum.max_row, column=4)
            cell.hyperlink = link
            cell.font = Font(color="0563C1", underline="single")

    # Sheet 2: Recipes
    ws_rec = wb.create_sheet("Recipes")
    rec_headers = [
        "Recipe ID",
        "Recipe Name",
        "Folder ID",
        "Running",
        "Central URL Properties",
        "Recipe Link",
        "Notes",
    ]
    ws_rec.append(rec_headers)
    for rr in sorted(recipe_rows, key=lambda x: x.get("properties", "")):
        ws_rec.append(
            [
                rr["recipe_id"],
                rr["recipe_name"],
                rr["folder_id"],
                rr["running"],
                rr["properties"],
                rr.get("recipe_link", ""),
                rr.get("note", ""),
            ]
        )
        if rr.get("recipe_link"):
            cell = ws_rec.cell(row=ws_rec.max_row, column=6)
            cell.hyperlink = rr["recipe_link"]
            cell.font = Font(color="0563C1", underline="single")

    # Sheet 3: Endpoints
    ws_ep = wb.create_sheet("API Endpoints")
    ep_headers = [
        "Account Property",
        "HTTP Method",
        "Endpoint Pattern",
        "Recipe ID",
        "Recipe Name",
        "Running",
        "Folder ID",
        "Recipe Link",
        "URL Template Snippet",
    ]
    ws_ep.append(ep_headers)
    for er in sorted(endpoint_rows, key=lambda x: (x["account_property"], x["recipe_id"])):
        ws_ep.append(
            [
                er["account_property"],
                er["http_method"],
                er["endpoint_pattern"],
                er["recipe_id"],
                er["recipe_name"],
                er["running"],
                er["folder_id"],
                er["recipe_link"],
                er["url_template_snippet"],
            ]
        )
        if er.get("recipe_link"):
            cell = ws_ep.cell(row=ws_ep.max_row, column=8)
            cell.hyperlink = er["recipe_link"]
            cell.font = Font(color="0563C1", underline="single")

    # Sheet 4: Connections
    ws_conn = wb.create_sheet("Central Connections")
    ws_conn.append(
        [
            "Connection ID",
            "Connection Name",
            "Application",
            "Auth Status",
            "Connection Link",
        ]
    )
    for cr in conn_rows:
        ws_conn.append(
            [
                cr.get("connection_id", ""),
                cr.get("connection_name", ""),
                cr.get("application", ""),
                cr.get("authorization_status", ""),
                cr.get("connection_link", ""),
            ]
        )
        if cr.get("connection_link"):
            cell = ws_conn.cell(row=ws_conn.max_row, column=5)
            cell.hyperlink = cr["connection_link"]
            cell.font = Font(color="0563C1", underline="single")

    # Sheet 5: Account property values (if API returned them)
    ws_prop = wb.create_sheet("Account Property Values")
    ws_prop.append(["Property Name", "Value (from API if available)", "Notes"])
    if prop_rows:
        for pr in prop_rows:
            ws_prop.append([pr.get("name", ""), pr.get("value", ""), ""])
    else:
        ws_prop.append(
            [
                "Central_url_v0",
                "(not exposed via Developer API — check Workato UI > Settings > Account properties)",
                "",
            ]
        )
        ws_prop.append(["Central_url_v2", "(see Workato UI)", ""])
        ws_prop.append(["Central_url_billing", "(see Workato UI)", ""])
        ws_prop.append(["Central_url_inventories", "(see Workato UI)", ""])

    for ws in wb.worksheets:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 50

    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")
    print(f"  Recipes with Central URLs: {len(recipe_rows)}")
    print(f"  Endpoint rows: {len(endpoint_rows)}")
    print(f"  Central connections: {len(conn_rows)}")


if __name__ == "__main__":
    main()
